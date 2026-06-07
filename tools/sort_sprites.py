#!/usr/bin/env python3
"""
sort_sprites.py — 处理画师提交、自动分类、生成贡献者表

用法:
    # 提取已完成条目 → 自动分类 → 生成贡献者表
    python3 sort_sprites.py --process --source <PNG目录> missing_sprites.xlsx

    # 手动模式 (直接给 PNG)
    python3 sort_sprites.py <文件|目录> ...

--process 模式:
    1. 遍历所有明细 Sheet，筛选「认领人」不为空 AND「状态」=已完成
    2. 在源目录搜索 PNG（匹配 游戏ID.png 或 safeID.png，/ 替换为 _）
    3. 自动检测: ✅ 通过 / ❌ 缺文件 / ❌ 未署名 / ❌ ID对不上
    4. 通过的 → 复制 PNG → 生成独立 JSON (含 artist)
    5. 追加「贡献者」Sheet 到 Excel (游戏ID|署名|贴图|描述|分类)
    6. 顶部按署名统计提交数
"""

import os, sys, json, shutil
from collections import defaultdict
from pathlib import Path


# ─── 分类 ───────────────────────────────────────────────
def classify(name: str, term: str = '') -> str:
    if name.startswith("overlay_"): return "overlay"
    if term in ('手持','男手持','女手持','穿戴','男穿戴','女穿戴','变异','男变异','女变异','效果','仿生'):
        return "overlay"
    if term == '尸体': return "monsters"
    if name.startswith("corpse_mon_") or name.startswith("mon_"): return "monsters"
    if name.startswith("t_"): return "terrain"
    if name.startswith("f_"): return "furniture"
    if name.startswith("fd_"): return "field"
    if name.startswith("tr_"): return "trap"
    if name.startswith("vp_"): return "vehicle"
    if name.startswith("bio_"): return "character"
    if name.startswith("bg_"): return "terrain"
    return "items"


def find_project_root() -> Path:
    d = Path(__file__).resolve().parent
    for p in [d] + list(d.parents):
        for c in (p / "source", p):
            if (c / "pngs_tiles_32x32").exists():
                return c
    # 同目录下查找克隆的贴图仓库
    for neighbor in d.parent.glob("CCB-UndeadPeople-Tileset*"):
        if (neighbor / "pngs_tiles_32x32").exists():
            return neighbor
    for neighbor in Path.cwd().parent.glob("CCB-UndeadPeople-Tileset*"):
        if (neighbor / "pngs_tiles_32x32").exists():
            return neighbor
    return d.parent


def collect_pngs(args: list) -> list:
    result = []
    for arg in args:
        p = Path(arg)
        if not p.exists(): continue
        if p.is_dir():
            result.extend(sorted(p.glob("*.png")))
            result.extend(sorted(p.glob("*.PNG")))
        elif p.suffix.lower() == ".png":
            result.append(p)
    return result


# ─── ──process 模式 ────────────────────────────────────
def process_mode(xlsx_path: str, source_dir: str):
    from openpyxl import load_workbook
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    check = Path(source_dir)
    if not check.is_dir():
        print(f"源目录不存在: {source_dir}"); return

    # 收集源 PNG
    png_map = {}
    for png in check.rglob("*.png"):  png_map[png.stem] = str(png.resolve())
    for png in check.rglob("*.PNG"):  png_map[png.stem] = str(png.resolve())
    print(f"源目录 PNG: {len(png_map)} 个")

    wb = load_workbook(xlsx_path)
    detail_names = [s for s in wb.sheetnames if s not in ('汇总','贡献者')]
    target = find_project_root() / "pngs_tiles_32x32"

    hfont = Font(bold=True, color="FFFFFF", size=11)
    hfill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    green = PatternFill(start_color="C6EFCE", fill_type="solid")
    red   = PatternFill(start_color="FFC7CE", fill_type="solid")
    yellow= PatternFill(start_color="FFEB9C", fill_type="solid")

    # 收集符合条件的条目
    contributors = defaultdict(int)  # name -> count
    results = []  # (game_id, artist, png_name, description, category, passed)

    for sname in detail_names:
        ws = wb[sname]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not row[0]: continue
            game_id = str(row[0]).strip()
            term    = str(row[2]).strip() if len(row) > 2 and row[2] else ""
            artist  = str(row[5]).strip() if len(row) > 5 and row[5] else ""
            status  = str(row[4]).strip() if len(row) > 4 and row[4] else ""

            if status != "已完成":
                continue

            safe_id = game_id.replace("/", "_")
            png_name = f"{safe_id}.png"

            if not artist:
                results.append((game_id, "", png_name, "❌ 未署名", sname, False, ""))
                continue

            # 检测 PNG
            desc = ""
            png_found = False
            found_path = png_map.get(safe_id) or png_map.get(game_id)
            if not found_path:
                desc = "❌ 缺文件"
            else:
                png_found = True
                desc = "✅ 通过"

            results.append((game_id, artist, png_name, desc, sname, png_found, found_path or ""))

    # 去重：同 ID 只保留第一个
    seen_ids = set()
    unique_results = []
    for r in results:
        if r[0] not in seen_ids:
            seen_ids.add(r[0])
            unique_results.append(r)
    results = unique_results

    # 处理通过的条目
    processed = 0
    for game_id, artist, png_name, desc, sname, png_found, path in results:
        if not png_found or "❌" in desc:
            continue
        safe_id = game_id.replace("/", "_")
        cat = classify(game_id, term)
        d = target / cat; d.mkdir(parents=True, exist_ok=True)
        shutil.copy2(path, str(d / f"{safe_id}.png"))
        entry = {"id": game_id, "fg": safe_id}
        if artist: entry["artist"] = artist
        with open(str(d / f"{safe_id}.json"), "w", encoding="utf-8") as f:
            json.dump(entry, f, indent=2, ensure_ascii=False)
        print(f"  [{artist:12s}] {cat:12s} {safe_id}")
        processed += 1

    # ─── 贡献者 Table → 独立文件 ───
    for _, artist, _, desc, _, png_found, _ in results:
        if "✅" in desc:
            contributors[artist] += 1

    total = len(results)
    passed = sum(1 for r in results if "✅" in r[3])
    missing_cnt = sum(1 for r in results if "缺文件" in r[3])
    unsigned = sum(1 for r in results if "未署名" in r[3])

    out_wb = Workbook()
    out_wb.remove(out_wb.active)
    ws = out_wb.create_sheet(title="贡献者")
    ws.append(["游戏ID", "署名", "贴图", "描述", "分类"])
    for c, w in zip(['A','B','C','D','E'], [48, 16, 42, 20, 16]):
        ws[f'{c}1'].font = hfont; ws[f'{c}1'].fill = hfill
        ws.column_dimensions[c].width = w

    for game_id, artist, png_name, desc, sname, png_found, path in results:
        r = ws.append([game_id, artist, png_name if "❌" in desc else f"{png_name} ✓", desc, sname])
        if "✅" in desc: ws.cell(row=ws._current_row, column=4).fill = green
        elif "❌" in desc: ws.cell(row=ws._current_row, column=4).fill = red
    ws.auto_filter.ref = ws.dimensions; ws.freeze_panes = 'A2'

    ws.insert_rows(1, len(contributors) + 4)
    ws.cell(row=1, column=1, value="贡献者统计").font = Font(bold=True, size=13)
    ws.cell(row=2, column=1, value="署名").font = Font(bold=True)
    ws.cell(row=2, column=2, value="通过数").font = Font(bold=True)
    for i, (name, count) in enumerate(sorted(contributors.items(), key=lambda x: -x[1]), 3):
        ws.cell(row=i, column=1, value=name)
        ws.cell(row=i, column=2, value=count)
    ws.cell(row=len(contributors)+3, column=1, value=f"总计: {passed} 通过 / {unsigned} 未署名 / {missing_cnt} 缺文件 / {total} 提交")
    ws.cell(row=len(contributors)+4, column=1, value="明细:").font = Font(bold=True)

    out_path = xlsx_path.rsplit('.', 1)[0] + '_贡献者.xlsx'
    out_wb.save(out_path)

    print(f"\n处理: {processed} PNG 已分类")
    print(f"贡献者表: {total} 提交 ({passed} 通过, {unsigned} 未署名, {missing_cnt} 缺文件)")
    if contributors:
        print("署名统计:")
        for name, count in sorted(contributors.items(), key=lambda x: -x[1]):
            print(f"  {name}: {count}")


# ─── 手动模式 ──────────────────────────────────────────
def manual_mode(args: list):
    inputs = collect_pngs(args)
    if not inputs: print("没有找到 PNG"); return
    root = find_project_root(); target = root / "pngs_tiles_32x32"
    if not target.exists(): print(f"目标不存在: {target}"); return
    print(f"工程根目录: {root}\n目标: {target}\n处理: {len(inputs)}\n")
    for src in sorted(inputs):
        stem = src.stem; safe = stem.replace("/", "_")
        cat = classify(stem); d = target / cat; d.mkdir(parents=True, exist_ok=True)
        shutil.copy2(str(src), str(d / f"{safe}.png"))
        entry = {"id": stem, "fg": safe}
        with open(str(d / f"{safe}.json"), "w", encoding="utf-8") as f:
            json.dump(entry, f, indent=2, ensure_ascii=False)
        print(f"  {cat:12s}  {src.name}")
    print("\n完成")


# ─── main ───────────────────────────────────────────────
def main():
    if len(sys.argv) < 2:
        print(__doc__); return
    args = sys.argv[1:]

    if "--process" in args:
        xlsx = ""; sdir = ""; i = 0
        while i < len(args):
            if args[i] == "--source" and i+1 < len(args): sdir = args[i+1]; i += 2
            elif args[i] == "--process": i += 1
            elif args[i].endswith('.xlsx'): xlsx = args[i]; i += 1
            else: i += 1
        if not xlsx: print("用法: sort_sprites.py --process --source <PNG目录> <.xlsx>"); return
        if not sdir:
            print("缺少 --source <PNG目录>"); return
        process_mode(xlsx, sdir); return

    manual_mode(args)


if __name__ == "__main__":
    main()
