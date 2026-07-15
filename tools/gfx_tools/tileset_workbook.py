#!/usr/bin/env python3
"""
生成 CCB 贴图覆盖率与贡献追踪工作簿。

脚本直接读取 CCB 的新物品格式（``type: ITEM`` + ``subtypes``），解析
``copy-from`` 继承，并区分：

* 已有：期望的贴图 ID 直接存在；
* 回退可用：通过 ``looks_like``、中性/性别贴图或兼容旧 ID 可以显示；
* 缺失：没有直接贴图，也没有可用回退。

默认扫描 ``data/json`` 和 ``data/mods``，跳过测试、调试与迁移数据。贴图源
仓库中使用时，通过 ``--data-dir`` 指向游戏仓库的数据目录。

用法：

  python3 tools/gfx_tools/compose.py --use-all --only-json . /tmp/composed
  python3 tools/gfx_tools/list_tileset_ids.py /tmp/composed > /tmp/tileset_ids.csv
  python3 tools/gfx_tools/tileset_workbook.py /tmp/tileset_ids.csv \
    --output 贴图贡献追踪表.xlsx

贴图源仓库中：

  python3 tools/gfx_tools/tileset_workbook.py /tmp/tileset_ids.csv \
    --data-dir ../Cataclysm-Cleanwater-Bomb/data/json \
    --data-dir ../Cataclysm-Cleanwater-Bomb/data/mods \
    --output 贴图贡献追踪表.xlsx
"""

import argparse
import csv
import glob
import json
import os

from dataclasses import dataclass


COLUMNS = ['类型', 'ID', '名称', '前缀', '完整贴图ID',
           '当前贴图', '完成状态', '提交', '贡献者', '备注']

SPRITE_HAVE = '已有'
SPRITE_FALLBACK = '回退可用'
SPRITE_MISSING = '缺失'
STATUS_MERGED = '已合并'
STATUS_TODO = '待领取'

DEFAULT_DATA_DIRS = ('data/json', 'data/mods')
EXCLUDED_DIR_NAMES = {'TEST_DATA', 'Standard_Combat_Tests'}

ITEM_WORN_SUBTYPES = {'ARMOR', 'TOOL_ARMOR'}

OVERMAP_TYPES = {'overmap_terrain', 'overmap_special'}
SIMPLE_TYPE_MAP = {
    'MONSTER': 'monsters',
    'terrain': 'terrain',
    'furniture': 'furniture',
    'mutation': 'mutation',
    'bionic': 'mutation',
    'vehicle_part': 'vehicle_part',
    'field_type': 'field',
    'trap': 'trap',
    'gate': 'gate',
    'SPELL': 'spell',
    'effect_type': 'effect',
    'movement_mode': 'movement_mode',
}

HARDCODED_IDS = (
    'cursor', 'highlight', 'highlight_item', 'footstep', 'graffiti',
    'zombie_revival_indicator', 'weather_rain_drop', 'weather_acid_drop',
    'weather_snowflake', 'animation_bullet_normal',
    'animation_bullet_shrapnel', 'animation_bullet_flame', 'explosion',
    'explosion_weak', 'explosion_medium', 'animation_hit', 'player_male',
    'player_female', 'npc_male', 'npc_female', 'animation_line',
    'line_target', 'line_trail', 'infrared_creature', 'run_nw', 'run_n',
    'run_ne', 'run_w', 'run_e', 'run_sw', 'run_s', 'run_se',
    'bash_complete', 'bash_effective', 'bash_ineffective', 'shadow',
    'overlay_hostile_sees_player', 'overlay_neutral_sees_player',
    'overlay_friendly_sees_player', 'overlay_other_sees_player',
)


@dataclass(frozen=True)
class Entity:
    gid: str
    name: str = ''
    looks_like: str = ''
    json_type: str = ''


def _as_list(value):
    if value is None:
        return []
    return value if isinstance(value, list) else [value]


def _object_refs(obj):
    refs = []
    abstract = obj.get('abstract')
    if isinstance(abstract, str):
        refs.append(abstract)
    refs.extend(x for x in _as_list(obj.get('id')) if isinstance(x, str))
    return refs


def _is_default_excluded(path):
    parts = os.path.normpath(path).split(os.sep)
    if any(part in EXCLUDED_DIR_NAMES for part in parts):
        return True
    if any(part.startswith('obsoletion_and_migration') for part in parts):
        return True
    return any(parts[i:i + 2] == ['json', 'debug']
               for i in range(max(0, len(parts) - 1)))


def _is_under(path, directory):
    path = os.path.abspath(path)
    directory = os.path.abspath(directory)
    try:
        return os.path.commonpath((path, directory)) == directory
    except ValueError:
        return False


def load_json_objects(data_dirs, exclude_dirs=()):
    """读取数据并返回 ``[(对象, 来源文件)]``；解析错误会明确失败。"""
    objects = []
    seen_files = set()
    for data_dir in data_dirs:
        for path in sorted(glob.glob(os.path.join(data_dir, '**', '*.json'),
                                     recursive=True)):
            real_path = os.path.realpath(path)
            if real_path in seen_files or _is_default_excluded(path):
                continue
            if any(_is_under(path, excluded) for excluded in exclude_dirs):
                continue
            seen_files.add(real_path)
            try:
                with open(path, encoding='utf-8') as source:
                    data = json.load(source)
            except Exception as err:
                raise RuntimeError(f'无法解析 JSON：{path}: {err}') from err
            if isinstance(data, dict):
                data = [data]
            if not isinstance(data, list):
                raise RuntimeError(f'JSON 顶层必须是对象或数组：{path}')
            for obj in data:
                if isinstance(obj, dict):
                    objects.append((obj, path))
    return objects


class JsonResolver:
    """解析覆盖率所需的少量 JSON 继承字段。"""

    SET_FIELDS = {'subtypes', 'flags'}

    def __init__(self, objects):
        self.by_ref = {}
        self.cache = {}
        for obj, _source in objects:
            for ref in _object_refs(obj):
                # 核心数据先加载；模组中的局部覆盖不应抹掉完整定义。
                self.by_ref.setdefault(ref, obj)

    def value(self, obj, field, trail=()):
        cache_key = (id(obj), field)
        if cache_key in self.cache:
            return self.cache[cache_key]

        parent_ref = obj.get('copy-from')
        parent = self.by_ref.get(parent_ref) if isinstance(parent_ref, str) else None
        parent_value = None
        if parent is not None and parent_ref not in trail:
            parent_value = self.value(parent, field, trail + (parent_ref,))

        if field in self.SET_FIELDS:
            if field in obj:
                result = set(_as_list(obj.get(field)))
            else:
                result = set(parent_value or ())
            extend = obj.get('extend')
            if isinstance(extend, dict):
                result.update(_as_list(extend.get(field)))
            delete = obj.get('delete')
            if isinstance(delete, dict):
                result.difference_update(_as_list(delete.get(field)))
            result = frozenset(x for x in result if isinstance(x, str))
        elif field in obj:
            result = obj.get(field)
        else:
            result = parent_value

        self.cache[cache_key] = result
        return result

    def looks_like_chain(self, entity):
        current = entity.looks_like
        seen = set()
        while isinstance(current, str) and current and current not in seen:
            seen.add(current)
            yield current
            target = self.by_ref.get(current)
            current = self.value(target, 'looks_like') if target else None


def get_name(value):
    if isinstance(value, dict):
        return value.get('str') or value.get('str_sp') or ''
    return str(value or '')


def load_csv_ids(path):
    with open(path, newline='', encoding='utf-8') as source:
        return [row[0].strip() for row in csv.reader(source)
                if row and row[0].strip()]


def scan_game_data(data_dirs, exclude_dirs=()):
    objects = load_json_objects(data_dirs, exclude_dirs)
    resolver = JsonResolver(objects)
    bucket_names = (
        'items', 'items_worn', 'items_wielded', 'monsters', 'terrain',
        'furniture', 'mutation', 'vehicle_part', 'field', 'trap', 'gate',
        'spell', 'effect', 'movement_mode', 'overmap', 'map_extra',
    )
    buckets = {name: [] for name in bucket_names}
    seen = {name: set() for name in bucket_names}

    def add(bucket, entity):
        if entity.gid and entity.gid not in seen[bucket]:
            seen[bucket].add(entity.gid)
            buckets[bucket].append(entity)

    for obj, _source in objects:
        if obj.get('abstract'):
            continue
        json_type = resolver.value(obj, 'type')
        if not isinstance(json_type, str):
            continue
        looks_like = resolver.value(obj, 'looks_like')
        looks_like = looks_like if isinstance(looks_like, str) else ''
        name = get_name(resolver.value(obj, 'name'))
        raw_ids = _as_list(obj.get('id'))
        if json_type in OVERMAP_TYPES and not raw_ids:
            raw_ids = _as_list(obj.get('om_terrain'))

        for gid in (x for x in raw_ids if isinstance(x, str)):
            entity = Entity(gid, name, looks_like, json_type)
            if json_type == 'ITEM':
                flags = resolver.value(obj, 'flags') or frozenset()
                subtypes = resolver.value(obj, 'subtypes') or frozenset()
                add('items', entity)
                overlay_allowed = (
                    not ({'PSEUDO', 'NO_DROP'} & set(flags)) and
                    obj.get('copy-from') not in ('fake_item', 'software')
                )
                if overlay_allowed and ITEM_WORN_SUBTYPES & set(subtypes):
                    add('items_worn', entity)
                # Character::get_overlay_ids() emits a wielded overlay for any
                # armed item, independently of its ITEM subtypes.
                if overlay_allowed:
                    add('items_wielded', entity)
            elif json_type in SIMPLE_TYPE_MAP:
                add(SIMPLE_TYPE_MAP[json_type], entity)
            elif json_type in OVERMAP_TYPES:
                add('overmap', entity)
            elif json_type == 'map_extra':
                add('map_extra', entity)

    return buckets, resolver


def sprite_state(tileset_ids, expected_id, fallback_ids=()):
    if expected_id in tileset_ids:
        return SPRITE_HAVE, expected_id
    for fallback_id in fallback_ids:
        if fallback_id and fallback_id in tileset_ids:
            return SPRITE_FALLBACK, fallback_id
    return SPRITE_MISSING, ''


def make_row(type_label, entity, prefix, full_id, state, resolved_id=''):
    note = f'回退贴图：{resolved_id}' if state == SPRITE_FALLBACK else ''
    return {
        '类型': type_label,
        'ID': entity.gid,
        '名称': entity.name,
        '前缀': prefix,
        '完整贴图ID': full_id,
        '当前贴图': state,
        '完成状态': STATUS_TODO if state == SPRITE_MISSING else STATUS_MERGED,
        '提交': '',
        '贡献者': '',
        '备注': note,
    }


def sort_rows(rows):
    order = {SPRITE_MISSING: 0, SPRITE_FALLBACK: 1, SPRITE_HAVE: 2}
    return sorted(rows, key=lambda row: (order[row['当前贴图']],
                                         row['类型'], row['ID']))


def prefixed_fallbacks(entity, resolver, prefix, include_raw=False):
    result = []
    for target in resolver.looks_like_chain(entity):
        result.append(target if target.startswith(prefix) else prefix + target)
        if include_raw:
            result.append(target)
    return result


def build_sheets(buckets, resolver, tileset_ids):
    ts = set(tileset_ids)
    sheets = {}

    def rows_for_entities(label, entities, prefix='', include_raw_fallback=False):
        rows = []
        for entity in entities:
            expected = prefix + entity.gid
            fallbacks = prefixed_fallbacks(
                entity, resolver, prefix, include_raw_fallback)
            state, found = sprite_state(ts, expected, fallbacks)
            rows.append(make_row(label, entity, prefix, expected, state, found))
        return sort_rows(rows)

    sheets['主贴图-物品'] = rows_for_entities('物品', buckets['items'])
    sheets['主贴图-怪物'] = rows_for_entities('MONSTER', buckets['monsters'])
    sheets['主贴图-地形'] = rows_for_entities('terrain', buckets['terrain'])
    sheets['主贴图-家具'] = rows_for_entities('furniture', buckets['furniture'])

    vehicle_rows = []
    for entity in buckets['vehicle_part']:
        expected = 'vp_' + entity.gid
        fallbacks = [entity.gid]
        fallbacks.extend(prefixed_fallbacks(entity, resolver, 'vp_', True))
        state, found = sprite_state(ts, expected, fallbacks)
        vehicle_rows.append(make_row(
            'vehicle_part', entity, 'vp_', expected, state, found))
    sheets['主贴图-载具部件'] = sort_rows(vehicle_rows)

    other_rows = []
    for bucket, label in (('field', 'field_type'), ('trap', 'trap'),
                          ('gate', 'gate'), ('spell', 'SPELL')):
        other_rows.extend(rows_for_entities(label, buckets[bucket]))
    sheets['主贴图-陷阱场效'] = sort_rows(other_rows)

    def gender_overlay_rows(entities, prefix, neutral_prefix, label):
        rows = []
        for entity in entities:
            expected = prefix + entity.gid
            fallbacks = [neutral_prefix + entity.gid]
            for target in resolver.looks_like_chain(entity):
                fallbacks.extend((prefix + target, neutral_prefix + target))
            state, found = sprite_state(ts, expected, fallbacks)
            rows.append(make_row(label, entity, prefix, expected, state, found))
        return sort_rows(rows)

    neutral_worn = []
    for entity in buckets['items_worn']:
        expected = 'overlay_worn_' + entity.gid
        fallbacks = prefixed_fallbacks(entity, resolver, 'overlay_worn_')
        male = 'overlay_male_worn_' + entity.gid
        female = 'overlay_female_worn_' + entity.gid
        state, found = sprite_state(ts, expected, fallbacks)
        if state == SPRITE_MISSING and male in ts and female in ts:
            state, found = SPRITE_FALLBACK, f'{male} + {female}'
        neutral_worn.append(make_row(
            '穿戴', entity, 'overlay_worn_', expected, state, found))
    sheets['穿戴-中性'] = sort_rows(neutral_worn)
    sheets['穿戴-男'] = gender_overlay_rows(
        buckets['items_worn'], 'overlay_male_worn_', 'overlay_worn_', '穿戴-男')
    sheets['穿戴-女'] = gender_overlay_rows(
        buckets['items_worn'], 'overlay_female_worn_', 'overlay_worn_', '穿戴-女')

    sheets['手持'] = rows_for_entities(
        '手持', buckets['items_wielded'], 'overlay_wielded_')
    corpse_rows = []
    for entity in buckets['monsters']:
        expected = 'corpse_' + entity.gid
        fallbacks = ['corpse']
        fallbacks.extend(prefixed_fallbacks(entity, resolver, 'corpse_'))
        state, found = sprite_state(ts, expected, fallbacks)
        corpse_rows.append(make_row(
            '尸体', entity, 'corpse_', expected, state, found))
    sheets['尸体'] = sort_rows(corpse_rows)
    sheets['手持尸体'] = rows_for_entities(
        '手持尸体', buckets['monsters'], 'overlay_wielded_corpse_')

    neutral_mutation = []
    for entity in buckets['mutation']:
        expected = 'overlay_mutation_' + entity.gid
        fallbacks = prefixed_fallbacks(entity, resolver, 'overlay_mutation_')
        male = 'overlay_male_mutation_' + entity.gid
        female = 'overlay_female_mutation_' + entity.gid
        state, found = sprite_state(ts, expected, fallbacks)
        if state == SPRITE_MISSING and male in ts and female in ts:
            state, found = SPRITE_FALLBACK, f'{male} + {female}'
        neutral_mutation.append(make_row(
            '变异生化', entity, 'overlay_mutation_', expected, state, found))
    sheets['变异生化-中性'] = sort_rows(neutral_mutation)
    sheets['变异生化-男'] = gender_overlay_rows(
        buckets['mutation'], 'overlay_male_mutation_',
        'overlay_mutation_', '变异生化-男')
    sheets['变异生化-女'] = gender_overlay_rows(
        buckets['mutation'], 'overlay_female_mutation_',
        'overlay_mutation_', '变异生化-女')

    sheets['效果叠加'] = rows_for_entities(
        'effect_type', buckets['effect'], 'overlay_effect_')
    sheets['移动模式叠加'] = rows_for_entities(
        'movement_mode', buckets['movement_mode'], 'overlay_')

    overmap_rows = []
    for entity in buckets['overmap']:
        direct = sorted(x for x in ts
                        if x == entity.gid or x.startswith(entity.gid + '_'))
        fallbacks = []
        for target in resolver.looks_like_chain(entity):
            fallbacks.extend(sorted(
                x for x in ts if x == target or x.startswith(target + '_')))
        if direct:
            state, found = SPRITE_HAVE, direct[0]
        else:
            state, found = sprite_state(ts, entity.gid, fallbacks)
        overmap_rows.append(make_row(
            entity.json_type, entity, '', entity.gid, state, found))
    sheets['overmap'] = sort_rows(overmap_rows)

    map_extra_rows = []
    for entity in buckets['map_extra']:
        expected = entity.gid if entity.gid.startswith('mx_') else 'mx_' + entity.gid
        state, found = sprite_state(ts, expected)
        map_extra_rows.append(make_row(
            'map_extra', entity, 'mx_', expected, state, found))
    sheets['地图事件'] = sort_rows(map_extra_rows)

    hardcoded_rows = []
    for tile_id in HARDCODED_IDS:
        entity = Entity(tile_id, json_type='hardcoded')
        state, found = sprite_state(ts, tile_id)
        hardcoded_rows.append(make_row(
            'hardcoded', entity, '', tile_id, state, found))
    sheets['硬编码'] = sort_rows(hardcoded_rows)
    return sheets


SHEET_ORDER = [
    '主贴图-物品', '主贴图-怪物', '主贴图-地形', '主贴图-家具',
    '主贴图-载具部件', '主贴图-陷阱场效',
    '穿戴-中性', '穿戴-男', '穿戴-女', '手持',
    '变异生化-中性', '变异生化-男', '变异生化-女',
    '尸体', '手持尸体', '效果叠加', '移动模式叠加',
    'overmap', '地图事件', '硬编码',
]


def write_workbook(sheets, output_path):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    workbook = Workbook()
    workbook.remove(workbook.active)
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill('solid', fgColor='4472C4')
    missing_fill = PatternFill('solid', fgColor='FFE699')
    fallback_fill = PatternFill('solid', fgColor='DDEBF7')
    center = Alignment(horizontal='center', vertical='center')

    summary = workbook.create_sheet('汇总')
    summary.append(['工作表', '总数', '直接已有', '回退可用', '真缺失', '有效覆盖率'])
    for cell in summary[1]:
        cell.font, cell.fill, cell.alignment = header_font, header_fill, center

    for name in SHEET_ORDER:
        rows = sheets.get(name, [])
        worksheet = workbook.create_sheet(name)
        worksheet.append(COLUMNS)
        for cell in worksheet[1]:
            cell.font, cell.fill, cell.alignment = header_font, header_fill, center
        worksheet.freeze_panes = 'A2'
        for row in rows:
            worksheet.append([row[column] for column in COLUMNS])
            fill = None
            if row['当前贴图'] == SPRITE_MISSING:
                fill = missing_fill
            elif row['当前贴图'] == SPRITE_FALLBACK:
                fill = fallback_fill
            if fill:
                for cell in worksheet[worksheet.max_row]:
                    cell.fill = fill
        for column, width in {
                'A': 16, 'B': 32, 'C': 24, 'D': 24, 'E': 42,
                'F': 11, 'G': 9, 'H': 24, 'I': 12, 'J': 48}.items():
            worksheet.column_dimensions[column].width = width

        total = len(rows)
        direct = sum(row['当前贴图'] == SPRITE_HAVE for row in rows)
        fallback = sum(row['当前贴图'] == SPRITE_FALLBACK for row in rows)
        missing = total - direct - fallback
        coverage = f'{(direct + fallback) / total * 100:.1f}%' if total else '—'
        summary.append([name, total, direct, fallback, missing, coverage])

    for column, width in {
            'A': 20, 'B': 10, 'C': 12, 'D': 12, 'E': 10, 'F': 14}.items():
        summary.column_dimensions[column].width = width
    workbook.save(output_path)


def main():
    parser = argparse.ArgumentParser(
        description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    parser.add_argument('tileset_ids_csv', help='list_tileset_ids.py 的输出')
    parser.add_argument(
        '--data-dir', action='append', dest='data_dirs',
        help='游戏数据目录，可重复；默认扫描 data/json 与 data/mods')
    parser.add_argument(
        '--exclude-dir', action='append', default=[],
        help='额外排除目录，可重复')
    parser.add_argument('--output', default='贴图贡献追踪表.xlsx')
    args = parser.parse_args()

    data_dirs = args.data_dirs or list(DEFAULT_DATA_DIRS)
    missing_dirs = [path for path in data_dirs if not os.path.isdir(path)]
    if missing_dirs:
        parser.error('数据目录不存在：' + ', '.join(missing_dirs))

    tileset_ids = load_csv_ids(args.tileset_ids_csv)
    buckets, resolver = scan_game_data(data_dirs, args.exclude_dir)
    sheets = build_sheets(buckets, resolver, tileset_ids)
    write_workbook(sheets, args.output)

    print(f'✓ 已生成 {args.output}')
    grand_total = grand_direct = grand_fallback = grand_missing = 0
    for name in SHEET_ORDER:
        rows = sheets.get(name, [])
        direct = sum(row['当前贴图'] == SPRITE_HAVE for row in rows)
        fallback = sum(row['当前贴图'] == SPRITE_FALLBACK for row in rows)
        missing = len(rows) - direct - fallback
        grand_total += len(rows)
        grand_direct += direct
        grand_fallback += fallback
        grand_missing += missing
        print(f'    {name}: {len(rows)} 条（直接 {direct}，回退 {fallback}，缺失 {missing}）')
    effective = ((grand_direct + grand_fallback) / grand_total * 100
                 if grand_total else 0)
    print(f'  合计 {grand_total} 条：直接 {grand_direct}，回退 {grand_fallback}，'
          f'真缺失 {grand_missing}，有效覆盖率 {effective:.1f}%')


if __name__ == '__main__':
    main()
