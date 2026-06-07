#!/usr/bin/env python3
import json, os, glob, sys
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

NEED_TYPES = (
    'ITEM','GUN','TOOL','ARMOR','COMESTIBLE','BOOK','GENERIC',
    'AMMO','MONSTER','terrain','furniture','trap','field_type',
    'mutation','vehicle_part','bionic','overmap_terrain','city_building',
    'map_extra','SPELL','profession','scenario','martial_art',
    'fault','proficiency','recipe_category','construction_category'
)
def extract_str(v):
    if isinstance(v, dict): return v.get('str_sp', v.get('str', str(v)))
    return str(v) if v else ''
def detail_cat(oid, tp):
    oid_l = oid.lower()
    if any(oid.startswith(p) for p in ('mon_','corpse_mon_')): return '怪物'
    if oid.startswith('t_'): return '地形'
    if oid.startswith('f_'): return '家具'
    if oid.startswith('fd_'): return '场地效果'
    if oid.startswith('tr_'): return '陷阱'
    if oid.startswith('vp_'): return '载具零件'
    if oid.startswith('bio_'): return '仿生插件'
    if oid.startswith('overlay_'): return '叠加层'
    if oid.startswith('bg_'): return '地形'
    if tp == 'MONSTER': return '怪物'
    if tp == 'terrain': return '地形'
    if tp == 'furniture': return '家具'
    if tp == 'field_type': return '场地效果'
    if tp == 'trap': return '陷阱'
    if tp == 'vehicle_part': return '载具零件'
    if tp == 'bionic': return '仿生插件'
    if tp == 'mutation': return '变异'
    if tp in ('overmap_terrain','city_building','map_extra'): return '地图建筑'
    if tp == 'SPELL': return '咒语'
    if tp == 'profession': return '职业'
    if tp == 'scenario': return '场景'
    if tp == 'martial_art': return '武术'
    if tp == 'fault': return '故障'
    if tp == 'proficiency': return '专精'
    if tp == 'recipe_category': return '配方类'
    if tp == 'construction_category': return '建筑类'
    for kw,cat in [('gun','武器弹药'),('rifle','武器弹药'),('pistol','武器弹药'),('shotgun','武器弹药'),
        ('smg','武器弹药'),('revolver','武器弹药'),('carbine','武器弹药'),('launcher','武器弹药'),
        ('ammo','武器弹药'),('mag','武器弹药'),('bullet','武器弹药'),('shell','武器弹药'),
        ('rocket','武器弹药'),('grenade','武器弹药'),('bomb','武器弹药'),('explosive','武器弹药'),
        ('bolt','武器弹药'),('arrow','武器弹药'),('crossbow','武器弹药'),
        ('sword','武器弹药'),('axe','武器弹药'),('knife','武器弹药'),('mace','武器弹药'),
        ('hammer','武器弹药'),('halberd','武器弹药'),('flail','武器弹药'),
        ('saw','工具'),('wrench','工具'),('drill','工具'),('screw','工具'),
        ('tool','工具'),('kit','工具'),('jack','工具'),('welder','工具'),
        ('armor','护甲'),('helmet','护甲'),('boots','护甲'),('gloves','护甲'),
        ('plate','护甲'),('cuirass','护甲'),('greave','护甲'),('gauntlet','护甲'),
        ('vest','护甲'),('mask','护甲'),('goggle','护甲'),('suit','护甲'),
        ('food','食物药品'),('drink','食物药品'),('meat','食物药品'),('fruit','食物药品'),
        ('vegetable','食物药品'),('bread','食物药品'),('soup','食物药品'),('pizza','食物药品'),
        ('drug','食物药品'),('pill','食物药品'),('bandage','食物药品'),('medicine','食物药品'),
        ('mutagen','食物药品'),('serum','食物药品'),('inject','食物药品'),('adrenaline','食物药品'),
        ('book','书籍'),('manual','书籍'),('recipe','书籍'),('scroll','书籍'),
        ('textbook','书籍'),('guide','书籍'),('magazine','书籍')]:
        if kw in oid_l: return cat
    return '材料杂项'

# 叠加层生成器
def gen_overlays(base_id, tp, name):
    base_name = name or base_id
    r = [(base_id, base_name, '基础', 'P0', tp)]
    if tp in ('ITEM','GUN','TOOL','ARMOR','COMESTIBLE','BOOK','GENERIC','AMMO'):
        r.extend([
            (f'overlay_wielded_{base_id}', f'手持 {base_name}', '手持', 'P0', tp),
            (f'overlay_male_wielded_{base_id}', f'男手持 {base_name}', '男手持', 'P1', tp),
            (f'overlay_female_wielded_{base_id}', f'女手持 {base_name}', '女手持', 'P1', tp),
            (f'overlay_worn_{base_id}', f'穿戴 {base_name}', '穿戴', 'P0', tp),
            (f'overlay_male_worn_{base_id}', f'男穿戴 {base_name}', '男穿戴', 'P1', tp),
            (f'overlay_female_worn_{base_id}', f'女穿戴 {base_name}', '女穿戴', 'P1', tp)])
    elif tp == 'MONSTER':
        r.append((f'corpse_mon_{base_id}', f'尸体 {base_name}', '尸体', 'P0', tp))
    elif tp == 'mutation':
        r.extend([
            (f'overlay_mutation_{base_id}', f'变异 {base_name}', '变异', 'P0', tp),
            (f'overlay_male_mutation_{base_id}', f'男变异 {base_name}', '男变异', 'P1', tp),
            (f'overlay_female_mutation_{base_id}', f'女变异 {base_name}', '女变异', 'P1', tp)])
    elif tp == 'bionic':
        r.append((f'overlay_wielded_bio_{base_id}', f'仿生 {base_name}', '仿生', 'P1', tp))
    elif tp == 'field_type':
        r.append((f'overlay_effect_{base_id}', f'效果 {base_name}', '效果', 'P1', tp))
    return r

print("收集 CCB 数据...")
id_info = {}
oid_names = {}
for f in glob.glob('data/json/**/*.json', recursive=True):
    if any(x in f for x in ('TEST_DATA','obsolete','obsoletion','mods/')): continue
    try:
        with open(f) as fh: data = json.load(fh)
    except: continue
    if not isinstance(data, list): data = [data]
    for o in data:
        if not isinstance(o, dict): continue
        tp = o.get('type',''); oid = o.get('id','')
        if isinstance(oid, list): oid = oid[0] if oid else ''
        if not oid or o.get('abstract'): continue
        if tp in NEED_TYPES:
            name = extract_str(o.get('name', ''))
            desc = str(o.get('description',''))
            if isinstance(o.get('description'), dict): desc = extract_str(o.get('description'))
            if len(desc) > 100: desc = desc[:97] + '...'
            ll = o.get('looks_like','')
            if isinstance(ll, dict): ll = ll.get('str','')
            if o.get('copy-from') and ll:
                desc = f'[复用 {ll}] ' + (desc if desc else '')
            id_info[oid] = {'name': name, 'desc': desc, 'type': tp}
            oid_names[oid] = {'name': name, 'desc': desc}

src = os.path.expanduser('~/Desktop/MSX++UnDeadPeople_source/source')
tids = set()
for jf in glob.glob(f'{src}/pngs_*/*/*.json'):
    with open(jf) as f:
        data = json.load(f)
    if not isinstance(data, list): data = [data]
    for e in data:
            eid = e.get('id','')
            if isinstance(eid, list): tids.update(eid)
            elif eid: tids.add(eid)

missing = set(id_info.keys()) - tids
print(f"CCB: {len(id_info)} IDs, 已有: {len(tids)}, 缺失: {len(missing)}")

all_rows = defaultdict(list)
for oid in sorted(missing):
    info = id_info.get(oid,{}); tp = info.get('type',''); name = info.get('name','')
    cat = detail_cat(oid, tp)
    desc = info.get('desc','')
    overlays = gen_overlays(oid, tp, name)
    for ov_id, ov_name, term, pri, t in overlays:
        all_rows[cat].append((ov_id, ov_name, term, pri, t, desc))

DETAIL_ORDER = [
    '武器弹药','工具','护甲','食物药品','书籍','材料杂项',
    '怪物','地形','家具','载具零件','场地效果','陷阱','地图建筑',
    '变异','仿生插件',
    '咒语','职业','场景','武术','故障','专精','配方类','建筑类',
]

wb = Workbook(); wb.remove(wb.active)
hfont = Font(bold=True, color="FFFFFF", size=11)
hfill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
gfill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

ws = wb.create_sheet(title='汇总', index=0)
for c,w in zip(['A','B','C','D','E'],[18,10,10,10,10]): ws.column_dimensions[c].width = w
ws.append(['类别','总数','已完成','进行中','待认领'])
for c in ['A','B','C','D','E']: ws[f'{c}1'].font = hfont; ws[f'{c}1'].fill = hfill
tr=2
for sname in DETAIL_ORDER:
    if sname not in all_rows: continue
    ws.append([sname,f"=COUNTA('{sname}'!A:A)-1",
        f"=COUNTIF('{sname}'!E:E,\"已完成\")",
        f"=COUNTIF('{sname}'!E:E,\"进行中\")",
        f"=COUNTIF('{sname}'!E:E,\"待画\")"]); tr+=1
ws.append(['总计',f"=SUM(B2:B{tr-1})",f"=SUM(C2:C{tr-1})",f"=SUM(D2:D{tr-1})",f"=SUM(E2:E{tr-1})"])
for c in ['A','B','C','D','E']: ws[f'{c}{tr}'].font = Font(bold=True)

total=0
for sname in DETAIL_ORDER:
    if sname not in all_rows or not all_rows[sname]: continue
    ws = wb.create_sheet(title=sname[:31])
    ws.append(['游戏ID','名称','词条','上传','状态','认领人','备注'])
    for c,w in zip(['A','B','C','D','E','F','G'],[52,30,10,10,10,14,40]):
        ws[f'{c}1'].font = hfont; ws[f'{c}1'].fill = hfill; ws.column_dimensions[c].width = w
    # P0 first, then P1
    for ov_id, ov_name, term, pri, t, desc in sorted(all_rows[sname], key=lambda x: (x[3]!='P0', x[0])):
        r = ws.append([ov_id, ov_name, term, '', '待画', '', desc if desc else ''])
        if pri == 'P1':
            for col_ord in range(1,8):
                ws.cell(row=ws._current_row, column=col_ord).fill = gfill
            ws.cell(row=ws._current_row, column=7).value = '非必须，基础完成后再画'
        total+=1
    ws.auto_filter.ref = ws.dimensions; ws.freeze_panes = 'A2'

out = os.path.expanduser('~/Desktop/missing_sprites.xlsx')
wb.save(out)
print(f"\nDone: {out}")
print(f"总计: {total} 行")
for sname in DETAIL_ORDER:
    if sname in all_rows: print(f"  {sname}: {len(all_rows[sname])}")
